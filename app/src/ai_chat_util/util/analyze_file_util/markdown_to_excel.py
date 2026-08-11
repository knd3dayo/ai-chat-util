from __future__ import annotations

from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from .markdown_parser import MarkdownParser


class MarkdownToExcelUtil:
    @classmethod
    def convert_markdown_to_excel(cls, markdown_text: str, output_path: str | Path) -> str:
        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        ws.freeze_panes = "A2"

        blocks = MarkdownParser.parse(markdown_text)
        rows: list[list[str]] = []
        for block in blocks:
            if block.type == "heading" and block.text:
                rows.append([f"Heading {block.level}", block.text])
            elif block.type == "paragraph" and block.text:
                rows.append(["Paragraph", block.text])
            elif block.type == "list_item" and block.text:
                rows.append(["List", block.text])
            elif block.type == "code_block" and block.code:
                rows.append(["Code", block.code])
            elif block.type == "table" and block.rows:
                rows.extend(block.rows)
            elif block.type == "image" and block.image_path:
                rows.append(["Image", block.image_path])

        for row_index, row in enumerate(rows, start=1):
            for col_index, value in enumerate(row, start=1):
                ws.cell(row=row_index, column=col_index, value=value)

        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        thin_border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in row:
                cell.border = thin_border
                cell.alignment = Alignment(vertical="top")

        output = Path(output_path)
        output.parent.mkdir(parents=True, exist_ok=True)
        wb.save(str(output))
        return str(output)
