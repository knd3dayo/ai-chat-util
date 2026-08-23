from pathlib import Path
import re

from openpyxl import Workbook
from openpyxl.styles import Border, Side
import pytest

from ai_chat_util.util.analyze_file_util.excel_util import ExcelUtil


THIN = Side(style="thin")


def _set_cell_border_side(ws, row: int, col: int, side_name: str) -> None:
    cell = ws.cell(row=row, column=col)
    current = cell.border
    kwargs = {
        "left": current.left,
        "right": current.right,
        "top": current.top,
        "bottom": current.bottom,
    }
    kwargs[side_name] = THIN
    cell.border = Border(**kwargs)


def _draw_perimeter_border(ws, start_row: int, start_col: int, end_row: int, end_col: int, with_bottom: bool = True) -> None:
    for col in range(start_col, end_col + 1):
        _set_cell_border_side(ws, start_row, col, "top")
        if with_bottom:
            _set_cell_border_side(ws, end_row, col, "bottom")

    for row in range(start_row, end_row + 1):
        _set_cell_border_side(ws, row, start_col, "left")
        _set_cell_border_side(ws, row, end_col, "right")


def _put_values(ws, start_row: int, start_col: int, values: list[list[object]]) -> None:
    for row_offset, row_values in enumerate(values):
        for col_offset, value in enumerate(row_values):
            ws.cell(row=start_row + row_offset, column=start_col + col_offset, value=value)


def _save_wb(tmp_path: Path, wb: Workbook, filename: str) -> str:
    path = tmp_path / filename
    wb.save(path)
    return str(path)


def _load_expected_ranges_from_answer(answer_path: Path) -> list[tuple[str, list[str]]]:
    expectations: list[tuple[str, list[str]]] = []
    current_file: str | None = None
    current_ranges: list[str] = []

    for raw_line in answer_path.read_text(encoding="utf-8").splitlines():
        line = raw_line.strip()
        if not line:
            continue

        if line.startswith("## "):
            if current_file is not None:
                expectations.append((current_file, current_ranges))
            title = line[3:].strip()
            current_file = title if title.endswith(".xlsx") else None
            current_ranges = []
            continue

        if current_file is None:
            continue

        matched = re.search(r"([A-Z]+\d+:[A-Z]+\d+)", line)
        if matched:
            current_ranges.append(matched.group(1))

    if current_file is not None:
        expectations.append((current_file, current_ranges))

    return expectations


def test_detect_single_table(tmp_path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    _draw_perimeter_border(ws, 2, 2, 4, 4)
    _put_values(ws, 2, 2, [["h1", "h2", "h3"], [1, 2, 3], [4, 5, 6]])

    filename = _save_wb(tmp_path, wb, "single.xlsx")
    tables = ExcelUtil.detect_tables_in_sheet(filename, sheet_name="Sheet1")

    assert len(tables) == 1
    assert tables[0]["a1_range"] == "B2:D4"


def test_detect_multiple_tables(tmp_path: Path) -> None:
    wb = Workbook()
    ws = wb.active

    _draw_perimeter_border(ws, 2, 2, 4, 4)
    _put_values(ws, 2, 2, [["a", "b", "c"], [1, 2, 3], [4, 5, 6]])

    _draw_perimeter_border(ws, 7, 2, 9, 3)
    _put_values(ws, 7, 2, [["x", "y"], [7, 8], [9, 10]])

    filename = _save_wb(tmp_path, wb, "multi.xlsx")
    tables = ExcelUtil.detect_tables_in_sheet(filename)

    assert [table["a1_range"] for table in tables] == ["B2:D4", "B7:C9"]


def test_detect_table_with_partial_empty_cells(tmp_path: Path) -> None:
    wb = Workbook()
    ws = wb.active

    _draw_perimeter_border(ws, 2, 2, 5, 4)
    _put_values(
        ws,
        2,
        2,
        [["h1", "h2", "h3"], [1, None, 3], [4, 5, None], [7, 8, 9]],
    )

    filename = _save_wb(tmp_path, wb, "partial-empty.xlsx")
    tables = ExcelUtil.detect_tables_in_sheet(filename)

    assert len(tables) == 1
    assert tables[0]["a1_range"] == "B2:D5"


def test_detect_table_without_bottom_border(tmp_path: Path) -> None:
    wb = Workbook()
    ws = wb.active

    _draw_perimeter_border(ws, 2, 2, 4, 4, with_bottom=False)
    _put_values(ws, 2, 2, [["h1", "h2", "h3"], [1, 2, 3], [4, 5, 6]])

    ws.cell(row=20, column=10, value="outside")

    filename = _save_wb(tmp_path, wb, "no-bottom.xlsx")
    tables = ExcelUtil.detect_tables_in_sheet(filename, empty_row_tolerance=2)

    assert len(tables) == 1
    assert tables[0]["a1_range"] == "B2:D4"


def test_detect_adjacent_tables_without_over_inclusion(tmp_path: Path) -> None:
    wb = Workbook()
    ws = wb.active

    _draw_perimeter_border(ws, 2, 2, 6, 4)
    _put_values(ws, 2, 2, [["a", "b", "c"], [1, 2, 3], [4, 5, 6], [7, 8, 9], [10, 11, 12]])

    _draw_perimeter_border(ws, 3, 6, 7, 8)
    _put_values(ws, 3, 6, [["x", "y", "z"], [13, 14, 15], [16, 17, 18], [19, 20, 21], [22, 23, 24]])

    filename = _save_wb(tmp_path, wb, "adjacent.xlsx")
    tables = ExcelUtil.detect_tables_in_sheet(filename)

    assert [table["a1_range"] for table in tables] == ["B2:D6", "F3:H7"]


@pytest.mark.parametrize(
    ("fixture_name", "expected_ranges"),
    _load_expected_ranges_from_answer(Path("work/detect_tables_test/answer.md")),
)
def test_detect_tables_with_real_fixtures_from_answer_md(
    fixture_name: str,
    expected_ranges: list[str],
) -> None:
    fixture = Path("work/detect_tables_test") / fixture_name
    assert fixture.exists()

    tables = ExcelUtil.detect_tables_in_sheet(str(fixture), sheet_name="Sheet1", empty_row_tolerance=2)

    assert [table["a1_range"] for table in tables] == expected_ranges
