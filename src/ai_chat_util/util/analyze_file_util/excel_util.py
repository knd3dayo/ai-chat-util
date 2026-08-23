import datetime
import openpyxl
from io import StringIO
from typing import Any
from openpyxl.utils import get_column_letter

class ExcelUtil:

    # application/vnd.openxmlformats-officedocument.spreadsheetml.sheetのファイルを読み込んで文字列として返す関数
    @classmethod
    def extract_text_from_sheet(cls, filename:str, sheet_name:str=""):
        # 出力用のストリームを作成
        output = StringIO()
        wb = openpyxl.load_workbook(filename)
        for sheet in wb:
            # シート名が指定されている場合はそのシートのみ処理
            if sheet_name and sheet.title != sheet_name:
                continue
            for row in sheet.iter_rows(values_only=True):
                # 1行分のデータを格納するリスト
                cells = []
                for cell in row:
                    # cell.valueがNoneの場合はcontinue
                    if cell is None:
                        continue
                    # cell.valueがdatetime.datetimeの場合はisoformat()で文字列に変換
                    if isinstance(cell, datetime.datetime):
                        cells.append(cell.isoformat())
                    else:
                        cells.append(str(cell))
                    
                output.write("\t".join(cells))
                output.write("\n")
        
        return output.getvalue()

    # excelのシート名一覧を取得する関数
    @classmethod
    def get_sheet_names(cls, filename):
        import openpyxl
        wb = openpyxl.load_workbook(filename)
        return wb.sheetnames

    # データをExcelファイルにエクスポートする関数
    @classmethod
    def export_data_to_excel(cls, data: dict[str, list], filename, sheet_name: str| None ="Sheet1"):
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        if ws is None:
            ws = wb.create_sheet()
        if sheet_name is not None:
            ws.title = sheet_name

        # ヘッダー行の追加
        headers = list(data.keys())
        ws.append(headers)
        # データ行の追加
        num_rows = len(next(iter(data.values()), []))
        for i in range(num_rows):
            row = []
            for header in headers:
                column_data = data.get(header, [])
                if i < len(column_data):
                    row.append(column_data[i])
                else:
                    row.append("")
            ws.append(row)
        
        wb.save(filename)

    # Excelファイルの内容を辞書型で取得する関数
    @classmethod
    def import_data_from_excel(cls, filename, sheet_name: str | None ="Sheet1") -> dict[str, list]:
        import openpyxl
        wb = openpyxl.load_workbook(filename)
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
        else:
            ws = wb.active
        if ws is None:
            return {}
        
        data: dict[str, list] = {}
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return data

        headers = rows[0]
        for header in headers:
            data[str(header)] = []
        for row in rows[1:]:
            for header, cell in zip(headers, row):
                data[str(header)].append(cell)
        
        return data

    @staticmethod
    def _has_border(cell: Any, side_name: str) -> bool:
        side = getattr(cell.border, side_name, None)
        return side is not None and side.style is not None

    @classmethod
    def _has_left_border(cls, cell: Any) -> bool:
        return cls._has_border(cell, "left")

    @classmethod
    def _has_top_border(cls, cell: Any) -> bool:
        return cls._has_border(cell, "top")

    @classmethod
    def _has_right_border(cls, cell: Any) -> bool:
        return cls._has_border(cell, "right")

    @classmethod
    def _is_effectively_empty_cell(cls, cell: Any) -> bool:
        if cell.value is not None:
            return False
        return not (
            cls._has_border(cell, "left")
            or cls._has_border(cell, "right")
            or cls._has_border(cell, "top")
            or cls._has_border(cell, "bottom")
        )

    @staticmethod
    def _is_in_range(row: int, col: int, area: tuple[int, int, int, int]) -> bool:
        start_row, start_col, end_row, end_col = area
        return start_row <= row <= end_row and start_col <= col <= end_col

    @classmethod
    def _is_in_discovered_ranges(
        cls,
        row: int,
        col: int,
        discovered: list[tuple[int, int, int, int]],
    ) -> bool:
        return any(cls._is_in_range(row, col, area) for area in discovered)

    @staticmethod
    def _row_all_none(ws: Any, row: int, min_col: int, max_col: int) -> bool:
        for col in range(min_col, max_col + 1):
            if ws.cell(row=row, column=col).value is not None:
                return False
        return True

    @staticmethod
    def _col_all_none(ws: Any, col: int, min_row: int, max_row: int) -> bool:
        for row in range(min_row, max_row + 1):
            if ws.cell(row=row, column=col).value is not None:
                return False
        return True

    @classmethod
    def _crop_bbox_by_values(
        cls,
        ws: Any,
        bbox: tuple[int, int, int, int],
    ) -> tuple[int, int, int, int] | None:
        min_row, min_col, max_row, max_col = bbox

        while min_row <= max_row and cls._row_all_none(ws, min_row, min_col, max_col):
            min_row += 1
        while min_row <= max_row and cls._row_all_none(ws, max_row, min_col, max_col):
            max_row -= 1
        while min_col <= max_col and cls._col_all_none(ws, min_col, min_row, max_row):
            min_col += 1
        while min_col <= max_col and cls._col_all_none(ws, max_col, min_row, max_row):
            max_col -= 1

        if min_row > max_row or min_col > max_col:
            return None
        return (min_row, min_col, max_row, max_col)

    @classmethod
    def _row_has_data_or_vertical_border(
        cls,
        ws: Any,
        row: int,
        start_col: int,
        end_col: int,
    ) -> bool:
        has_vertical_border = False
        has_value_with_border = False
        for col in range(start_col, end_col + 1):
            cell = ws.cell(row=row, column=col)
            has_left_or_right = cls._has_border(cell, "left") or cls._has_border(cell, "right")
            if has_left_or_right:
                has_vertical_border = True

            if cell.value is not None and (
                has_left_or_right
                or cls._has_border(cell, "top")
                or cls._has_border(cell, "bottom")
            ):
                has_value_with_border = True

        return has_vertical_border or has_value_with_border

    @classmethod
    def _to_a1_range(
        cls,
        start_row: int,
        start_col: int,
        end_row: int,
        end_col: int,
    ) -> str:
        return (
            f"{get_column_letter(start_col)}{start_row}:"
            f"{get_column_letter(end_col)}{end_row}"
        )

    @classmethod
    def _is_anchor_cell(
        cls,
        ws: Any,
        row: int,
        col: int,
        min_row: int,
        min_col: int,
    ) -> bool:
        cell = ws.cell(row=row, column=col)
        if not (cls._has_left_border(cell) and cls._has_top_border(cell)):
            return False

        # 同じ罫線グリッド上の内側セルを起点にしないため、
        # 左隣に上罫線がある場合は左端ではないとみなす。
        if col > min_col:
            left_cell = ws.cell(row=row, column=col - 1)
            if cls._has_top_border(left_cell):
                return False

        # 同様に、上隣に左罫線がある場合は上端ではないとみなす。
        if row > min_row:
            top_cell = ws.cell(row=row - 1, column=col)
            if cls._has_left_border(top_cell):
                return False

        return True

    @classmethod
    def detect_tables_in_sheet(
        cls,
        filename: str,
        sheet_name: str | None = None,
        empty_row_tolerance: int = 2,
        scan_bbox: tuple[int, int, int, int] | None = None,
    ) -> list[dict[str, Any]]:
        wb = openpyxl.load_workbook(filename)
        ws = wb[sheet_name] if sheet_name and sheet_name in wb.sheetnames else wb.active
        if ws is None:
            return []

        tolerance = max(1, int(empty_row_tolerance))

        initial_bbox = scan_bbox or (ws.min_row, ws.min_column, ws.max_row, ws.max_column)
        cropped_bbox = cls._crop_bbox_by_values(ws, initial_bbox)
        if cropped_bbox is None:
            return []

        min_row, min_col, max_row, max_col = cropped_bbox
        discovered_ranges: list[tuple[int, int, int, int]] = []
        results: list[dict[str, Any]] = []

        for row in range(min_row, max_row + 1):
            for col in range(min_col, max_col + 1):
                if cls._is_in_discovered_ranges(row, col, discovered_ranges):
                    continue

                if not cls._is_anchor_cell(ws, row, col, min_row, min_col):
                    continue

                right_col = col
                for current_col in range(col, max_col + 1):
                    current_cell = ws.cell(row=row, column=current_col)
                    right_col = current_col

                    if cls._has_right_border(current_cell):
                        if current_col >= max_col:
                            break
                        next_cell = ws.cell(row=row, column=current_col + 1)
                        if (
                            not cls._is_effectively_empty_cell(next_cell)
                            and (
                                cls._has_left_border(next_cell)
                                or cls._has_top_border(next_cell)
                                or next_cell.value is not None
                            )
                        ):
                            continue
                        break

                    if current_col < max_col:
                        next_cell = ws.cell(row=row, column=current_col + 1)
                        if cls._is_effectively_empty_cell(next_cell):
                            break

                last_valid_row = row
                empty_streak = 0
                for current_row in range(row, max_row + 1):
                    if cls._row_has_data_or_vertical_border(ws, current_row, col, right_col):
                        last_valid_row = current_row
                        empty_streak = 0
                    else:
                        empty_streak += 1
                        if empty_streak >= tolerance:
                            break

                area = (row, col, last_valid_row, right_col)
                discovered_ranges.append(area)
                results.append(
                    {
                        "sheet_name": ws.title,
                        "start_row": row,
                        "start_col": col,
                        "end_row": last_valid_row,
                        "end_col": right_col,
                        "a1_range": cls._to_a1_range(row, col, last_valid_row, right_col),
                    }
                )

        return results